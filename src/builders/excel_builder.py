import os
import json
import logging
from datetime import datetime
import openpyxl

logging.basicConfig(level=logging.INFO, format='%(levelname)s - %(message)s')

class ExcelReportBuilder:
    def __init__(self, base_dir: str):
        self.base_dir = base_dir

    def _load_json(self, file_name: str):
        path = os.path.join(self.base_dir, file_name)
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logging.error(f"Error cargando JSON {file_name}: {e}")
            return None

    def build_report(self, template_name: str, blueprint_name: str, data_name: str, output_name: str):
        template_path = os.path.join(self.base_dir, template_name)
        output_path = os.path.join(self.base_dir, output_name)
        
        # 1. Cargar el plano y los datos
        blueprint = self._load_json(blueprint_name)
        data = self._load_json(data_name)
        
        if not blueprint or not data:
            logging.error("No se pudo cargar el Blueprint o los Datos.")
            return
            
        logging.info(f"Cargando plantilla: {template_name}...")
        try:
            # 2. Cargar Plantilla (preservando formatos y macros)
            # data_only=False preserva las fórmulas en lugar de evaluarlas a valores estáticos
            wb = openpyxl.load_workbook(template_path, data_only=False)
            
            # 3. Iterar sobre la estructura definida en el Blueprint
            for sheet_info in blueprint.get("estructura", []):
                sheet_name = sheet_info.get("hoja")
                elementos = sheet_info.get("elementos", [])
                
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    logging.info(f"Procesando hoja: '{sheet_name}'")
                    
                    for elemento in elementos:
                        logging.info(f" -> Localizado componente: {elemento}")
                        # AQUÍ VA LA LÓGICA DE INYECCIÓN
                        # Por ejemplo, si es "tabla_Tabla1", se busca la tabla y se inyecta data["participacion_porcentual"]...
                        
                        # Demo de inserción segura (agregamos una marca de tiempo para comprobar que escribimos)
                        # Buscamos la primera celda vacía en la columna A para no sobrescribir nada importante
                        max_row = sheet.max_row
                        sheet.cell(row=max_row + 2, column=1, value=f"Reporte Actualizado por SIEA Engine: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                else:
                    logging.warning(f"La hoja '{sheet_name}' definida en el Blueprint no existe en la Plantilla.")
            
            # 4. Guardar archivo final
            wb.save(output_path)
            logging.info(f"¡Reporte construido exitosamente! Guardado como: {output_name}")
            
        except Exception as e:
            logging.error(f"Error construyendo el reporte de Excel: {e}")

if __name__ == "__main__":
    base_dir = r"C:\Users\LENOVO X1 YOGA\Videos\AUTOMATIZACION\INFORMES QUE HAGO"
    
    builder = ExcelReportBuilder(base_dir)
    builder.build_report(
        template_name="Plantilla_SIEA.xlsx",
        blueprint_name="excel_blueprint.json",
        data_name="statistical_truth.json",
        output_name="Reporte_Final.xlsx"
    )
