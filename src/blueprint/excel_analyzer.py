import os
import logging
import openpyxl

class ExcelAnalyzer:
    """
    ExcelAnalyzer es responsable de extraer las hojas visibles y su metadata interna
    de una plantilla de Excel (.xlsx), como tablas, gráficos, fórmulas, colores, 
    protecciones y nombres de rangos globales. No modifica nada.
    """
    def analyze(self, file_path: str) -> dict:
        excel_data = {
            "template": os.path.basename(file_path),
            "worksheets": [],
            "metrics": {
                "named_ranges": [],
                "sheets_data": {}
            }
        }
        
        try:
            # openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=False) # data_only=False para poder ver fórmulas
            
            # 1. Extraer nombres de rangos definidos globales
            if hasattr(wb, 'defined_names'):
                for dn in wb.defined_names.definedName:
                    excel_data["metrics"]["named_ranges"].append(dn.name)
            
            # 2. Iterar sobre las hojas
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Ignorar hojas ocultas
                if sheet.sheet_state != 'visible':
                    continue
                
                excel_data["worksheets"].append(sheet_name)
                
                tables_count = len(sheet.tables) if hasattr(sheet, 'tables') else 0
                charts_count = len(sheet._charts) if hasattr(sheet, '_charts') else 0
                is_protected = sheet.protection.sheet if hasattr(sheet.protection, 'sheet') else False
                
                # Estructura base para esta hoja
                sheet_metrics = {
                    "tables": tables_count,
                    "charts": charts_count,
                    "formulas_detected": False,
                    "is_protected": is_protected,
                    "columns": [],
                    "sampled_colors": []
                }
                
                # Muestreo de las primeras filas
                colors_set = set()
                # Extraemos nombres de columnas de la fila 1 (si existen)
                for cell in sheet[1]:
                    if cell.value is not None:
                        sheet_metrics["columns"].append(str(cell.value).strip())
                
                # Muestreo para fórmulas y colores (máx 50 filas, para rendimiento)
                max_sample_rows = 50
                row_count = 0
                for row in sheet.iter_rows(min_row=1, max_row=max_sample_rows):
                    row_count += 1
                    for cell in row:
                        # Fórmulas
                        if not sheet_metrics["formulas_detected"]:
                            if isinstance(cell.value, str) and cell.value.startswith('='):
                                sheet_metrics["formulas_detected"] = True
                        
                        # Colores (Solid Fills)
                        if getattr(cell, 'fill', None) and cell.fill.fill_type == "solid":
                            color_val = getattr(cell.fill.start_color, 'rgb', None)
                            # Si start_color.rgb está disponible
                            if color_val and isinstance(color_val, str) and color_val != "00000000":
                                colors_set.add(color_val)
                                
                sheet_metrics["sampled_colors"] = list(colors_set)
                
                excel_data["metrics"]["sheets_data"][sheet_name] = sheet_metrics
                
            return excel_data
            
        except Exception as e:
            logging.error(f"Error al analizar el archivo Excel {file_path}: {e}")
            raise
