"""Tests para ExcelReader."""
import pytest
from pathlib import Path
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from src.ingestion.excel_reader import ExcelReader
from src.ingestion.exceptions import FileNotReadableError


@pytest.fixture
def valid_excel_file(tmp_path):
    """Fixture que crea un archivo Excel válido temporal."""
    file_path = tmp_path / "test_data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws["A1"] = "Test"
    wb.save(file_path)
    return file_path


@pytest.fixture
def invalid_extension_file(tmp_path):
    """Fixture que crea un archivo de texto con extensión no soportada."""
    file_path = tmp_path / "test_data.txt"
    file_path.write_text("dummy content")
    return file_path


@pytest.fixture
def corrupt_excel_file(tmp_path):
    """Fixture que crea un archivo .xlsx corrupto."""
    file_path = tmp_path / "corrupt.xlsx"
    file_path.write_bytes(b"This is not a valid zip or excel file")
    return file_path


def test_load_valid_workbook(valid_excel_file):
    """Verifica que se carga correctamente un workbook válido."""
    reader = ExcelReader()
    handle = reader.load_workbook(valid_excel_file)
    
    assert handle is not None
    assert handle.workbook is not None
    assert handle.source_path == valid_excel_file
    assert handle.readonly is True
    assert len(handle.checksum) == 64  # SHA-256 es de 64 caracteres hex
    
    # Verificamos que se haya cargado una hoja
    assert "Datos" in handle.workbook.sheetnames
    
    handle.workbook.close()


def test_load_non_existent_file(tmp_path):
    """Verifica que se levanta FileNotReadableError si el archivo no existe."""
    reader = ExcelReader()
    missing_file = tmp_path / "does_not_exist.xlsx"
    
    with pytest.raises(FileNotReadableError) as exc_info:
        reader.load_workbook(missing_file)
        
    assert "El archivo no existe" in str(exc_info.value)


def test_load_not_a_file(tmp_path):
    """Verifica que se levanta FileNotReadableError si la ruta es un directorio."""
    reader = ExcelReader()
    
    with pytest.raises(FileNotReadableError) as exc_info:
        reader.load_workbook(tmp_path)
        
    assert "La ruta no es un archivo válido" in str(exc_info.value)


def test_load_invalid_extension(invalid_extension_file):
    """Verifica que se levanta FileNotReadableError para extensiones no válidas."""
    reader = ExcelReader()
    
    with pytest.raises(FileNotReadableError) as exc_info:
        reader.load_workbook(invalid_extension_file)
        
    assert "Formato no soportado" in str(exc_info.value)


def test_load_corrupt_file(corrupt_excel_file):
    """Verifica que se levanta FileNotReadableError si el archivo está dañado."""
    reader = ExcelReader()
    
    with pytest.raises(FileNotReadableError) as exc_info:
        reader.load_workbook(corrupt_excel_file)
        
    assert "Archivo dañado o protegido con contraseña" in str(exc_info.value)


def test_unexpected_error(monkeypatch, valid_excel_file):
    """Verifica el manejo de excepciones inesperadas."""
    
    # Simulamos una excepción inesperada en load_workbook de openpyxl
    def mock_load(*args, **kwargs):
        raise RuntimeError("Error catastrófico")
        
    monkeypatch.setattr(openpyxl, "load_workbook", mock_load)
    
    reader = ExcelReader()
    
    with pytest.raises(FileNotReadableError) as exc_info:
        reader.load_workbook(valid_excel_file)
        
    assert "Error inesperado" in str(exc_info.value)
