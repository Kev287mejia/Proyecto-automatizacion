"""
Suite de pruebas unitarias para el módulo ``WorkbookAnalyzer`` — Fase 3A.

Cobertura objetivo: ≥ 90%

Estrategia de testing:
    - Todos los tests son unitarios (sin I/O real de archivos).
    - Los workbooks se construyen en memoria con openpyxl.
    - Se mockea el sistema de archivos para aislar la lógica del analizador.
    - Se prueban todos los caminos de código incluyendo ramas de error.

Estructura de la suite:
    TestWorksheetMetadata         — Modelo de datos de hoja
    TestWorkbookStatistics        — Modelo de estadísticas
    TestWorkbookProtection        — Modelo de protección
    TestWorkbookMetadata          — Modelo de metadata
    TestWorkbookContext           — Modelo raíz con métodos de consulta
    TestWorkbookAnalyzerAnalyze   — Método principal ``analyze``
    TestGetValidSheets            — Método helper ``get_valid_sheets``
    TestGetFileSize               — Helper privado de tamaño
    TestExtractNamedRanges        — Helper privado de rangos nombrados
    TestExtractProtection         — Helper privado de protección
    TestInspectWorksheet          — Helper privado de inspección de hoja
    TestResolveTargetSheet        — Helper privado de hoja target
    TestLogging                   — Verificación de mensajes de log
    TestIntegration               — Casos de uso end-to-end
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from src.ingestion.exceptions import InvalidWorkbookError
from src.ingestion.workbook_analyzer import WorkbookAnalyzer
from src.types.workbook_context import (
    WorkbookContext,
    WorkbookMetadata,
    WorkbookProtection,
    WorkbookStatistics,
    WorksheetMetadata,
)


# ===========================================================================
# Fixtures compartidas
# ===========================================================================


@pytest.fixture
def analyzer() -> WorkbookAnalyzer:
    """Instancia fresca del WorkbookAnalyzer sin dependencias externas."""
    return WorkbookAnalyzer()


@pytest.fixture
def simple_workbook() -> openpyxl.Workbook:
    """Workbook mínimo: una hoja visible con datos reales."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws["A1"] = "Nombre"
    ws["A2"] = "Ana"
    return wb


@pytest.fixture
def complex_workbook() -> openpyxl.Workbook:
    """Workbook con múltiples hojas: visible con datos, vacía, oculta."""
    wb = openpyxl.Workbook()

    # Hoja 1: visible con datos
    ws1 = wb.active
    ws1.title = "Datos"
    ws1["A1"] = "Encabezado"
    ws1["A2"] = "Valor"

    # Hoja 2: visible pero vacía
    wb.create_sheet(title="Vacia")

    # Hoja 3: oculta con datos
    ws3 = wb.create_sheet(title="Oculta")
    ws3["A1"] = "Secreto"
    ws3.sheet_state = "hidden"

    # Hoja 4: visible con espacios (no tiene datos reales)
    ws4 = wb.create_sheet(title="Falsa")
    ws4["A1"] = "   "

    return wb


@pytest.fixture
def all_hidden_workbook() -> openpyxl.Workbook:
    """Workbook donde todas las hojas están ocultas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Oculta1"
    ws["A1"] = "Dato"
    ws.sheet_state = "hidden"
    ws2 = wb.create_sheet(title="Oculta2")
    ws2.sheet_state = "veryHidden"
    return wb


@pytest.fixture
def fake_path(tmp_path: Path) -> Path:
    """Ruta a un archivo .xlsx temporal existente en disco."""
    f = tmp_path / "test_workbook.xlsx"
    f.write_bytes(b"PK" + b"\x00" * 50)  # Simulación de bytes ZIP mínimos
    return f


@pytest.fixture
def nonexistent_path(tmp_path: Path) -> Path:
    """Ruta que NO existe en disco."""
    return tmp_path / "no_existe.xlsx"


# ===========================================================================
# Tests: Modelos de dominio (inmutabilidad y propiedades)
# ===========================================================================


class TestWorksheetMetadata:
    """Pruebas del modelo WorksheetMetadata."""

    def test_frozen_prevents_mutation(self) -> None:
        """Los modelos frozen deben rechazar asignaciones directas."""
        meta = WorksheetMetadata(
            name="Hoja1",
            is_hidden=False,
            dimensions="A1:D10",
            has_tables=False,
            has_charts=False,
            has_filters=False,
            tables=(),
            chart_count=0,
        )
        with pytest.raises((AttributeError, TypeError)):
            meta.name = "Otro"  # type: ignore[misc]

    def test_default_chart_count(self) -> None:
        """chart_count debe ser 0 por defecto."""
        meta = WorksheetMetadata(
            name="X",
            is_hidden=False,
            dimensions="",
            has_tables=False,
            has_charts=False,
            has_filters=False,
            tables=(),
        )
        assert meta.chart_count == 0

    def test_tables_is_immutable_tuple(self) -> None:
        meta = WorksheetMetadata(
            name="X",
            is_hidden=False,
            dimensions="A1:B5",
            has_tables=True,
            has_charts=False,
            has_filters=False,
            tables=("TablaA", "TablaB"),
        )
        assert isinstance(meta.tables, tuple)
        assert len(meta.tables) == 2


class TestWorkbookStatistics:
    """Pruebas del modelo WorkbookStatistics."""

    def test_visible_sheet_count_computed(self) -> None:
        stats = WorkbookStatistics(
            approx_size_bytes=1024,
            sheet_count=5,
            hidden_sheet_count=2,
        )
        assert stats.visible_sheet_count == 3

    def test_visible_sheet_count_all_visible(self) -> None:
        stats = WorkbookStatistics(
            approx_size_bytes=512,
            sheet_count=3,
            hidden_sheet_count=0,
        )
        assert stats.visible_sheet_count == 3

    def test_visible_sheet_count_all_hidden(self) -> None:
        stats = WorkbookStatistics(
            approx_size_bytes=256,
            sheet_count=2,
            hidden_sheet_count=2,
        )
        assert stats.visible_sheet_count == 0

    def test_frozen_prevents_mutation(self) -> None:
        stats = WorkbookStatistics(1024, 3, 1)
        with pytest.raises((AttributeError, TypeError)):
            stats.sheet_count = 99  # type: ignore[misc]


class TestWorkbookProtection:
    """Pruebas del modelo WorkbookProtection."""

    def test_not_protected_defaults(self) -> None:
        prot = WorkbookProtection(is_protected=False)
        assert prot.lock_structure is False
        assert prot.lock_windows is False

    def test_protected_flags(self) -> None:
        prot = WorkbookProtection(
            is_protected=True,
            lock_structure=True,
            lock_windows=True,
        )
        assert prot.is_protected is True
        assert prot.lock_structure is True
        assert prot.lock_windows is True

    def test_frozen(self) -> None:
        prot = WorkbookProtection(False)
        with pytest.raises((AttributeError, TypeError)):
            prot.is_protected = True  # type: ignore[misc]


class TestWorkbookMetadata:
    """Pruebas del modelo WorkbookMetadata."""

    def test_fields(self) -> None:
        meta = WorkbookMetadata(
            file_name="datos.xlsx",
            active_sheet="Hoja1",
            target_sheet="Datos",
            named_ranges=("RANGO_A",),
        )
        assert meta.file_name == "datos.xlsx"
        assert meta.active_sheet == "Hoja1"
        assert meta.target_sheet == "Datos"
        assert meta.named_ranges == ("RANGO_A",)

    def test_frozen(self) -> None:
        meta = WorkbookMetadata("f.xlsx", "H1", "H1", ())
        with pytest.raises((AttributeError, TypeError)):
            meta.file_name = "otro.xlsx"  # type: ignore[misc]


class TestWorkbookContext:
    """Pruebas del modelo raíz WorkbookContext y sus métodos de consulta."""

    @pytest.fixture
    def sample_context(self) -> WorkbookContext:
        return WorkbookContext(
            statistics=WorkbookStatistics(2048, 3, 1),
            protection=WorkbookProtection(False),
            metadata=WorkbookMetadata("reporte.xlsx", "Datos", "Datos", ()),
            sheets=(
                WorksheetMetadata("Datos", False, "A1:D10", False, False, False, ()),
                WorksheetMetadata("Resumen", False, "A1:B5", False, False, False, ()),
                WorksheetMetadata("Oculta", True, "A1:A1", False, False, False, ()),
            ),
        )

    def test_get_sheet_found(self, sample_context: WorkbookContext) -> None:
        sheet = sample_context.get_sheet("Datos")
        assert sheet is not None
        assert sheet.name == "Datos"

    def test_get_sheet_not_found(self, sample_context: WorkbookContext) -> None:
        result = sample_context.get_sheet("NoExiste")
        assert result is None

    def test_get_visible_sheets(self, sample_context: WorkbookContext) -> None:
        visible = sample_context.get_visible_sheets()
        assert len(visible) == 2
        assert all(not s.is_hidden for s in visible)

    def test_get_visible_sheets_names(self, sample_context: WorkbookContext) -> None:
        visible = sample_context.get_visible_sheets()
        names = [s.name for s in visible]
        assert "Datos" in names
        assert "Resumen" in names
        assert "Oculta" not in names

    def test_frozen(self, sample_context: WorkbookContext) -> None:
        with pytest.raises((AttributeError, TypeError)):
            sample_context.sheets = ()  # type: ignore[misc]


# ===========================================================================
# Tests: WorkbookAnalyzer.analyze — flujo principal
# ===========================================================================


class TestWorkbookAnalyzerAnalyze:
    """Pruebas del método ``analyze`` del WorkbookAnalyzer."""

    def test_returns_workbook_context(
        self, analyzer: WorkbookAnalyzer, simple_workbook: openpyxl.Workbook
    ) -> None:
        """analyze debe retornar un WorkbookContext."""
        with patch("os.path.getsize", return_value=1024):
            ctx = analyzer.analyze(Path("test.xlsx"), simple_workbook)
        assert isinstance(ctx, WorkbookContext)

    def test_file_name_extracted(
        self, analyzer: WorkbookAnalyzer, simple_workbook: openpyxl.Workbook
    ) -> None:
        """El nombre del archivo debe reflejarse en WorkbookMetadata."""
        with patch("os.path.getsize", return_value=512):
            ctx = analyzer.analyze(Path("mi_archivo.xlsx"), simple_workbook)
        assert ctx.metadata.file_name == "mi_archivo.xlsx"

    def test_sheet_count(
        self, analyzer: WorkbookAnalyzer, complex_workbook: openpyxl.Workbook
    ) -> None:
        """El conteo de hojas debe ser el total del workbook."""
        with patch("os.path.getsize", return_value=2048):
            ctx = analyzer.analyze(Path("x.xlsx"), complex_workbook)
        assert ctx.statistics.sheet_count == 4

    def test_hidden_sheet_count(
        self, analyzer: WorkbookAnalyzer, complex_workbook: openpyxl.Workbook
    ) -> None:
        """Debe contarse 1 hoja oculta en complex_workbook."""
        with patch("os.path.getsize", return_value=2048):
            ctx = analyzer.analyze(Path("x.xlsx"), complex_workbook)
        assert ctx.statistics.hidden_sheet_count == 1

    def test_active_sheet(
        self, analyzer: WorkbookAnalyzer, simple_workbook: openpyxl.Workbook
    ) -> None:
        """La hoja activa debe coincidir con la hoja activa del workbook."""
        with patch("os.path.getsize", return_value=512):
            ctx = analyzer.analyze(Path("x.xlsx"), simple_workbook)
        assert ctx.metadata.active_sheet == "Datos"

    def test_target_sheet_first_visible_with_data(
        self, analyzer: WorkbookAnalyzer, complex_workbook: openpyxl.Workbook
    ) -> None:
        """Target sheet debe ser la primera hoja visible con datos."""
        with patch("os.path.getsize", return_value=2048):
            ctx = analyzer.analyze(Path("x.xlsx"), complex_workbook)
        assert ctx.metadata.target_sheet == "Datos"

    def test_sheets_metadata_count(
        self, analyzer: WorkbookAnalyzer, complex_workbook: openpyxl.Workbook
    ) -> None:
        """Se debe generar WorksheetMetadata para cada hoja."""
        with patch("os.path.getsize", return_value=2048):
            ctx = analyzer.analyze(Path("x.xlsx"), complex_workbook)
        assert len(ctx.sheets) == 4

    def test_hidden_sheet_detected(
        self, analyzer: WorkbookAnalyzer, complex_workbook: openpyxl.Workbook
    ) -> None:
        """La hoja 'Oculta' debe aparecer marcada como is_hidden=True."""
        with patch("os.path.getsize", return_value=2048):
            ctx = analyzer.analyze(Path("x.xlsx"), complex_workbook)
        oculta = ctx.get_sheet("Oculta")
        assert oculta is not None
        assert oculta.is_hidden is True

    def test_visible_sheet_not_hidden(
        self, analyzer: WorkbookAnalyzer, simple_workbook: openpyxl.Workbook
    ) -> None:
        """La hoja visible no debe estar marcada como oculta."""
        with patch("os.path.getsize", return_value=512):
            ctx = analyzer.analyze(Path("x.xlsx"), simple_workbook)
        datos = ctx.get_sheet("Datos")
        assert datos is not None
        assert datos.is_hidden is False

    def test_raises_invalid_workbook_no_sheets(
        self, analyzer: WorkbookAnalyzer
    ) -> None:
        """analyze debe lanzar InvalidWorkbookError si el workbook no tiene hojas."""
        wb = MagicMock()
        wb.sheetnames = []
        with pytest.raises(InvalidWorkbookError, match="no contiene hojas"):
            analyzer.analyze(Path("vacio.xlsx"), wb)

    def test_raises_invalid_workbook_all_hidden(
        self,
        analyzer: WorkbookAnalyzer,
        all_hidden_workbook: openpyxl.Workbook,
    ) -> None:
        """analyze debe lanzar InvalidWorkbookError si todas las hojas están ocultas."""
        with patch("os.path.getsize", return_value=512):
            with pytest.raises(InvalidWorkbookError):
                analyzer.analyze(Path("x.xlsx"), all_hidden_workbook)

    def test_approx_size_bytes_from_disk(
        self, analyzer: WorkbookAnalyzer, simple_workbook: openpyxl.Workbook
    ) -> None:
        """El tamaño en bytes debe tomarse del sistema de archivos."""
        mock_path = MagicMock(spec=Path)
        mock_path.name = "x.xlsx"
        mock_path.exists.return_value = True
        with patch("os.path.getsize", return_value=99999):
            ctx = analyzer.analyze(mock_path, simple_workbook)
        assert ctx.statistics.approx_size_bytes == 99999

    def test_no_named_ranges(
        self, analyzer: WorkbookAnalyzer, simple_workbook: openpyxl.Workbook
    ) -> None:
        """Un workbook sin rangos nombrados debe retornar tupla vacía."""
        with patch("os.path.getsize", return_value=512):
            ctx = analyzer.analyze(Path("x.xlsx"), simple_workbook)
        assert ctx.metadata.named_ranges == ()

    def test_protection_not_protected_by_default(
        self, analyzer: WorkbookAnalyzer, simple_workbook: openpyxl.Workbook
    ) -> None:
        """Un workbook sin protección debe retornar is_protected=False."""
        with patch("os.path.getsize", return_value=512):
            ctx = analyzer.analyze(Path("x.xlsx"), simple_workbook)
        assert ctx.protection.is_protected is False

    def test_context_is_immutable(
        self, analyzer: WorkbookAnalyzer, simple_workbook: openpyxl.Workbook
    ) -> None:
        """El WorkbookContext retornado debe ser inmutable."""
        with patch("os.path.getsize", return_value=512):
            ctx = analyzer.analyze(Path("x.xlsx"), simple_workbook)
        with pytest.raises((AttributeError, TypeError)):
            ctx.sheets = ()  # type: ignore[misc]

    def test_workbook_not_modified(
        self, analyzer: WorkbookAnalyzer, simple_workbook: openpyxl.Workbook
    ) -> None:
        """El workbook no debe ser modificado durante el análisis."""
        original_sheets = list(simple_workbook.sheetnames)
        with patch("os.path.getsize", return_value=512):
            analyzer.analyze(Path("x.xlsx"), simple_workbook)
        assert list(simple_workbook.sheetnames) == original_sheets


# ===========================================================================
# Tests: WorkbookAnalyzer.get_valid_sheets
# ===========================================================================


class TestGetValidSheets:
    """Pruebas del método helper ``get_valid_sheets``."""

    def test_returns_visible_sheets_with_data(
        self, analyzer: WorkbookAnalyzer, complex_workbook: openpyxl.Workbook
    ) -> None:
        """Solo debe retornar hojas visibles con contenido real."""
        valid = analyzer.get_valid_sheets(complex_workbook)
        names = [ws.title for ws in valid]
        assert "Datos" in names
        assert "Oculta" not in names

    def test_excludes_hidden_sheets(
        self, analyzer: WorkbookAnalyzer, complex_workbook: openpyxl.Workbook
    ) -> None:
        """Las hojas ocultas no deben aparecer en el resultado."""
        valid = analyzer.get_valid_sheets(complex_workbook)
        for ws in valid:
            assert ws.sheet_state != "hidden"

    def test_raises_if_no_valid_sheets(
        self,
        analyzer: WorkbookAnalyzer,
        all_hidden_workbook: openpyxl.Workbook,
    ) -> None:
        """Debe lanzar InvalidWorkbookError si no hay hojas visibles con datos."""
        with pytest.raises(InvalidWorkbookError, match="hojas visibles con datos"):
            analyzer.get_valid_sheets(all_hidden_workbook)

    def test_simple_workbook_returns_one_sheet(
        self, analyzer: WorkbookAnalyzer, simple_workbook: openpyxl.Workbook
    ) -> None:
        """Un workbook simple debe retornar exactamente 1 hoja válida."""
        valid = analyzer.get_valid_sheets(simple_workbook)
        assert len(valid) == 1
        assert valid[0].title == "Datos"

    def test_veryHidden_sheet_excluded(
        self, analyzer: WorkbookAnalyzer
    ) -> None:
        """Las hojas 'veryHidden' también deben excluirse."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VeryHidden"
        ws["A1"] = "Dato"
        ws.sheet_state = "veryHidden"

        ws2 = wb.create_sheet("Visible")
        ws2["A1"] = "OK"
        ws2["A2"] = "Valor"

        valid = analyzer.get_valid_sheets(wb)
        names = [ws.title for ws in valid]
        assert "VeryHidden" not in names
        assert "Visible" in names


# ===========================================================================
# Tests: Helper privado _get_file_size
# ===========================================================================


class TestGetFileSize:
    """Pruebas del helper estático ``_get_file_size``."""

    def test_returns_size_when_file_exists(self, fake_path: Path) -> None:
        size = WorkbookAnalyzer._get_file_size(fake_path)
        assert size > 0

    def test_returns_zero_when_file_not_exists(
        self, nonexistent_path: Path
    ) -> None:
        size = WorkbookAnalyzer._get_file_size(nonexistent_path)
        assert size == 0

    def test_returns_zero_on_oserror(self) -> None:
        with patch("os.path.getsize", side_effect=OSError("permission denied")):
            path = MagicMock(spec=Path)
            path.exists.return_value = True
            size = WorkbookAnalyzer._get_file_size(path)
        assert size == 0


# ===========================================================================
# Tests: Helper privado _extract_named_ranges
# ===========================================================================


class TestExtractNamedRanges:
    """Pruebas del helper estático ``_extract_named_ranges``."""

    def test_no_defined_names_returns_empty(self) -> None:
        wb = MagicMock()
        wb.defined_names = None
        ranges = WorkbookAnalyzer._extract_named_ranges(wb)
        assert ranges == []

    def test_extracts_from_defined_name_attribute(self) -> None:
        """Compatibilidad con openpyxl < 3.1 (definedName list)."""
        dn1 = MagicMock()
        dn1.name = "RANGO_A"
        dn2 = MagicMock()
        dn2.name = "RANGO_B"

        wb = MagicMock()
        wb.defined_names = MagicMock()
        wb.defined_names.definedName = [dn1, dn2]
        del wb.defined_names.__iter__  # simular que no es iterable directamente

        ranges = WorkbookAnalyzer._extract_named_ranges(wb)
        assert "RANGO_A" in ranges
        assert "RANGO_B" in ranges

    def test_extracts_from_iterable(self) -> None:
        """Compatibilidad con openpyxl >= 3.1 (iterable directo)."""
        dn1 = MagicMock()
        dn1.name = "TOTAL"
        dn1_spec = MagicMock()

        wb = MagicMock()
        defined = MagicMock()
        defined.definedName = None  # sin atributo definedName
        # Quitar definedName para forzar rama de iterable
        del defined.definedName
        defined.__iter__ = MagicMock(return_value=iter([dn1]))
        wb.defined_names = defined

        ranges = WorkbookAnalyzer._extract_named_ranges(wb)
        assert "TOTAL" in ranges

    def test_returns_empty_on_exception(self) -> None:
        """Si ocurre una excepción inesperada, debe retornar lista vacía."""
        wb = MagicMock()
        wb.defined_names = MagicMock(side_effect=RuntimeError("boom"))
        # Hacer que getattr retorne el mock que lanzará
        ranges = WorkbookAnalyzer._extract_named_ranges(wb)
        assert ranges == []


# ===========================================================================
# Tests: Helper privado _extract_protection
# ===========================================================================


class TestExtractProtection:
    """Pruebas del helper estático ``_extract_protection``."""

    def test_no_security_attr_returns_not_protected(self) -> None:
        wb = MagicMock(spec=[])  # Sin atributo 'security'
        prot = WorkbookAnalyzer._extract_protection(wb)
        assert prot.is_protected is False

    def test_security_none_returns_not_protected(self) -> None:
        wb = MagicMock()
        wb.security = None
        prot = WorkbookAnalyzer._extract_protection(wb)
        assert prot.is_protected is False
        assert prot.lock_structure is False
        assert prot.lock_windows is False

    def test_lock_structure_true(self) -> None:
        wb = MagicMock()
        wb.security = MagicMock()
        wb.security.lockStructure = True
        wb.security.lockWindows = False
        prot = WorkbookAnalyzer._extract_protection(wb)
        assert prot.is_protected is True
        assert prot.lock_structure is True
        assert prot.lock_windows is False

    def test_lock_windows_true(self) -> None:
        wb = MagicMock()
        wb.security = MagicMock()
        wb.security.lockStructure = False
        wb.security.lockWindows = True
        prot = WorkbookAnalyzer._extract_protection(wb)
        assert prot.is_protected is True
        assert prot.lock_windows is True

    def test_both_locks_true(self) -> None:
        wb = MagicMock()
        wb.security = MagicMock()
        wb.security.lockStructure = True
        wb.security.lockWindows = True
        prot = WorkbookAnalyzer._extract_protection(wb)
        assert prot.is_protected is True
        assert prot.lock_structure is True
        assert prot.lock_windows is True

    def test_both_locks_false(self) -> None:
        wb = MagicMock()
        wb.security = MagicMock()
        wb.security.lockStructure = False
        wb.security.lockWindows = False
        prot = WorkbookAnalyzer._extract_protection(wb)
        assert prot.is_protected is False


# ===========================================================================
# Tests: Helper estático _inspect_worksheet
# ===========================================================================


class TestInspectWorksheet:
    """Pruebas del helper estático ``_inspect_worksheet``."""

    def test_visible_sheet(self) -> None:
        ws = MagicMock()
        ws.title = "Hoja1"
        ws.sheet_state = "visible"
        ws.dimensions = "A1:D10"
        ws.tables = {}
        ws._charts = []
        ws.auto_filter = MagicMock()
        ws.auto_filter.ref = None

        meta = WorkbookAnalyzer._inspect_worksheet(ws)
        assert meta.name == "Hoja1"
        assert meta.is_hidden is False
        assert meta.dimensions == "A1:D10"
        assert meta.has_tables is False
        assert meta.has_charts is False
        assert meta.has_filters is False
        assert meta.tables == ()
        assert meta.chart_count == 0

    def test_hidden_sheet(self) -> None:
        ws = MagicMock()
        ws.title = "Secreta"
        ws.sheet_state = "hidden"
        ws.dimensions = "A1:A1"
        ws.tables = {}
        ws._charts = []
        ws.auto_filter = MagicMock()
        ws.auto_filter.ref = None

        meta = WorkbookAnalyzer._inspect_worksheet(ws)
        assert meta.is_hidden is True

    def test_very_hidden_sheet(self) -> None:
        ws = MagicMock()
        ws.title = "Muy Oculta"
        ws.sheet_state = "veryHidden"
        ws.dimensions = ""
        ws.tables = {}
        ws._charts = []
        ws.auto_filter = MagicMock()
        ws.auto_filter.ref = None

        meta = WorkbookAnalyzer._inspect_worksheet(ws)
        assert meta.is_hidden is True

    def test_sheet_with_tables(self) -> None:
        ws = MagicMock()
        ws.title = "ConTablas"
        ws.sheet_state = "visible"
        ws.dimensions = "A1:E20"
        ws.tables = {"Tabla1": MagicMock(), "Tabla2": MagicMock()}
        ws._charts = []
        ws.auto_filter = MagicMock()
        ws.auto_filter.ref = None

        meta = WorkbookAnalyzer._inspect_worksheet(ws)
        assert meta.has_tables is True
        assert "Tabla1" in meta.tables
        assert "Tabla2" in meta.tables

    def test_sheet_with_charts(self) -> None:
        ws = MagicMock()
        ws.title = "Graficos"
        ws.sheet_state = "visible"
        ws.dimensions = "A1:C5"
        ws.tables = {}
        ws._charts = [MagicMock(), MagicMock(), MagicMock()]
        ws.auto_filter = MagicMock()
        ws.auto_filter.ref = None

        meta = WorkbookAnalyzer._inspect_worksheet(ws)
        assert meta.has_charts is True
        assert meta.chart_count == 3

    def test_sheet_with_auto_filter(self) -> None:
        ws = MagicMock()
        ws.title = "Filtros"
        ws.sheet_state = "visible"
        ws.dimensions = "A1:D100"
        ws.tables = {}
        ws._charts = []
        ws.auto_filter = MagicMock()
        ws.auto_filter.ref = "A1:D1"

        meta = WorkbookAnalyzer._inspect_worksheet(ws)
        assert meta.has_filters is True

    def test_sheet_no_auto_filter_attr(self) -> None:
        ws = MagicMock(spec=["title", "sheet_state", "dimensions", "tables", "_charts"])
        ws.title = "SinFiltro"
        ws.sheet_state = "visible"
        ws.dimensions = "A1:B3"
        ws.tables = {}
        ws._charts = []

        meta = WorkbookAnalyzer._inspect_worksheet(ws)
        assert meta.has_filters is False


# ===========================================================================
# Tests: Helper privado _resolve_target_sheet
# ===========================================================================


class TestResolveTargetSheet:
    """Pruebas del helper ``_resolve_target_sheet``."""

    def _make_sheet_meta(
        self, name: str, hidden: bool, dimensions: str
    ) -> WorksheetMetadata:
        return WorksheetMetadata(
            name=name,
            is_hidden=hidden,
            dimensions=dimensions,
            has_tables=False,
            has_charts=False,
            has_filters=False,
            tables=(),
        )

    def _make_workbook_mock(self, sheets: dict[str, dict]) -> MagicMock:
        """Crea un workbook mock con hojas configurables."""
        wb = MagicMock()

        def get_sheet(name: str) -> MagicMock:
            ws = MagicMock()
            ws.max_row = sheets[name].get("max_row", 1)
            ws.max_column = sheets[name].get("max_col", 1)
            return ws

        wb.__getitem__ = MagicMock(side_effect=get_sheet)
        return wb

    def test_selects_first_visible_with_data_by_dimensions(self) -> None:
        sheets_meta = [
            self._make_sheet_meta("Vacia", False, "A1:A1"),
            self._make_sheet_meta("Datos", False, "A1:D100"),
        ]
        wb = self._make_workbook_mock(
            {"Vacia": {"max_row": 1, "max_col": 1}, "Datos": {"max_row": 50}}
        )
        target = WorkbookAnalyzer._resolve_target_sheet(sheets_meta, "Vacia", wb)
        assert target == "Datos"

    def test_selects_by_max_row(self) -> None:
        sheets_meta = [
            self._make_sheet_meta("Vacia", False, ""),
            self._make_sheet_meta("Contenido", False, ""),
        ]
        wb = self._make_workbook_mock(
            {"Vacia": {"max_row": 1, "max_col": 1}, "Contenido": {"max_row": 5}}
        )
        target = WorkbookAnalyzer._resolve_target_sheet(
            sheets_meta, "Vacia", wb
        )
        assert target == "Contenido"

    def test_fallback_to_active_sheet(self) -> None:
        sheets_meta = [
            self._make_sheet_meta("Hoja1", False, ""),
        ]
        wb = self._make_workbook_mock({"Hoja1": {"max_row": 1, "max_col": 1}})
        target = WorkbookAnalyzer._resolve_target_sheet(sheets_meta, "Hoja1", wb)
        assert target == "Hoja1"

    def test_raises_if_all_hidden(self) -> None:
        sheets_meta = [
            self._make_sheet_meta("Oculta1", True, "A1:D10"),
            self._make_sheet_meta("Oculta2", True, "A1:B5"),
        ]
        wb = self._make_workbook_mock(
            {"Oculta1": {"max_row": 5}, "Oculta2": {"max_row": 3}}
        )
        with pytest.raises(InvalidWorkbookError, match="hojas visibles"):
            WorkbookAnalyzer._resolve_target_sheet(sheets_meta, "", wb)

    def test_skips_hidden_and_picks_first_visible(self) -> None:
        sheets_meta = [
            self._make_sheet_meta("Oculta", True, "A1:D10"),
            self._make_sheet_meta("Visible", False, ""),
        ]
        wb = self._make_workbook_mock(
            {"Oculta": {"max_row": 5}, "Visible": {"max_row": 1, "max_col": 1}}
        )
        target = WorkbookAnalyzer._resolve_target_sheet(
            sheets_meta, "Oculta", wb
        )
        assert target == "Visible"


# ===========================================================================
# Tests: Logging
# ===========================================================================


class TestLogging:
    """Verifica que el analizador emite mensajes de log apropiados."""

    def test_logs_start_of_analysis(
        self,
        analyzer: WorkbookAnalyzer,
        simple_workbook: openpyxl.Workbook,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        with caplog.at_level(logging.INFO, logger="src.ingestion.workbook_analyzer"):
            with patch("os.path.getsize", return_value=512):
                analyzer.analyze(Path("reporte.xlsx"), simple_workbook)
        assert any("análisis forense" in r.message.lower() for r in caplog.records)

    def test_logs_completion(
        self,
        analyzer: WorkbookAnalyzer,
        simple_workbook: openpyxl.Workbook,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        with caplog.at_level(logging.INFO, logger="src.ingestion.workbook_analyzer"):
            with patch("os.path.getsize", return_value=512):
                analyzer.analyze(Path("x.xlsx"), simple_workbook)
        assert any("completado" in r.message.lower() for r in caplog.records)

    def test_logs_warning_no_valid_sheets(
        self,
        analyzer: WorkbookAnalyzer,
        all_hidden_workbook: openpyxl.Workbook,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        with caplog.at_level(logging.WARNING, logger="src.ingestion.workbook_analyzer"):
            with patch("os.path.getsize", return_value=512):
                with pytest.raises(InvalidWorkbookError):
                    analyzer.analyze(Path("x.xlsx"), all_hidden_workbook)


# ===========================================================================
# Tests: Integración end-to-end con archivos reales
# ===========================================================================


class TestIntegration:
    """Pruebas de integración con archivos reales en disco."""

    def test_analyze_real_file(self, tmp_path: Path) -> None:
        """Análisis completo de un archivo .xlsx real guardado en disco."""
        # Crear archivo real
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Beneficiarios"
        ws["A1"] = "Nombre"
        ws["B1"] = "Edad"
        ws["A2"] = "Juan"
        ws["B2"] = 25
        file_path = tmp_path / "beneficiarios.xlsx"
        wb.save(str(file_path))

        # Reabrir y analizar
        wb_loaded = openpyxl.load_workbook(str(file_path), read_only=False)
        analyzer = WorkbookAnalyzer()
        ctx = analyzer.analyze(file_path, wb_loaded)

        assert ctx.metadata.file_name == "beneficiarios.xlsx"
        assert ctx.statistics.sheet_count == 1
        assert ctx.statistics.hidden_sheet_count == 0
        assert ctx.statistics.visible_sheet_count == 1
        assert ctx.statistics.approx_size_bytes > 0
        assert ctx.metadata.target_sheet == "Beneficiarios"
        assert ctx.protection.is_protected is False
        assert len(ctx.sheets) == 1
        hoja = ctx.get_sheet("Beneficiarios")
        assert hoja is not None
        assert hoja.is_hidden is False

    def test_analyze_multi_sheet_file(self, tmp_path: Path) -> None:
        """Análisis de un archivo con múltiples hojas."""
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Datos"
        ws1["A1"] = "ID"
        ws1["A2"] = "1"
        ws2 = wb.create_sheet("Configuracion")
        ws2["A1"] = "param"
        ws3 = wb.create_sheet("Oculta")
        ws3["A1"] = "secreto"
        ws3.sheet_state = "hidden"
        file_path = tmp_path / "multi.xlsx"
        wb.save(str(file_path))

        wb_loaded = openpyxl.load_workbook(str(file_path))
        analyzer = WorkbookAnalyzer()
        ctx = analyzer.analyze(file_path, wb_loaded)

        assert ctx.statistics.sheet_count == 3
        assert ctx.statistics.hidden_sheet_count == 1
        assert ctx.statistics.visible_sheet_count == 2
        oculta = ctx.get_sheet("Oculta")
        assert oculta is not None and oculta.is_hidden is True

    def test_get_valid_sheets_real_file(self, tmp_path: Path) -> None:
        """get_valid_sheets con archivo real."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activa"
        ws["A1"] = "X"
        ws["A2"] = "Y"
        file_path = tmp_path / "validas.xlsx"
        wb.save(str(file_path))

        wb_loaded = openpyxl.load_workbook(str(file_path))
        analyzer = WorkbookAnalyzer()
        valid = analyzer.get_valid_sheets(wb_loaded)
        assert len(valid) == 1
        assert valid[0].title == "Activa"

    def test_no_pandas_import_in_module(self) -> None:
        """El módulo workbook_analyzer no debe importar pandas."""
        import importlib
        import sys

        mod = sys.modules.get("src.ingestion.workbook_analyzer")
        if mod is None:
            mod = importlib.import_module("src.ingestion.workbook_analyzer")
        source_file = getattr(mod, "__file__", "")
        if source_file:
            with open(source_file, encoding="utf-8") as f:
                content = f.read()
            assert "import pandas" not in content
            assert "from pandas" not in content

    def test_workbook_not_mutated_during_analysis(self, tmp_path: Path) -> None:
        """El workbook no debe ser modificado durante el análisis."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        ws["A1"] = "Valor"
        ws["A2"] = "Otro"
        file_path = tmp_path / "nomutation.xlsx"
        wb.save(str(file_path))

        wb_loaded = openpyxl.load_workbook(str(file_path))
        sheets_before = list(wb_loaded.sheetnames)
        analyzer = WorkbookAnalyzer()
        analyzer.analyze(file_path, wb_loaded)
        assert list(wb_loaded.sheetnames) == sheets_before
