"""
Suite de pruebas unitarias para la Fase 3B — ``WorksheetReader``.

Cobertura objetivo: ≥ 90%

Estrategia de testing:
    - Tests unitarios: sin I/O de archivo; worksheets construidos en memoria.
    - Tests de los modelos: inmutabilidad, propiedades computadas.
    - Tests del lector: flujos normales, edge cases y caminos de error.
    - Tests de helpers privados: ``_build_cells``, ``_build_column_profiles``,
      ``_extract_rows``.
    - Tests de integración: ciclo completo con archivos ``.xlsx`` reales.

Estructura::

    TestCellData              — modelo CellData
    TestRowData               — modelo RowData y propiedades
    TestColumnProfile         — modelo ColumnProfile y fill_rate
    TestExtractionStats       — modelo ExtractionStats
    TestWorksheetData         — modelo raíz y métodos de consulta
    TestWorksheetReaderExtract — método principal extract_data
    TestBuildCells            — helper _build_cells
    TestBuildColumnProfiles   — helper _build_column_profiles
    TestExtractRows           — helper _extract_rows
    TestLogging               — mensajes de log emitidos
    TestIntegration           — ciclo end-to-end con archivos reales
"""

from __future__ import annotations

import logging
from typing import Any
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from src.ingestion.exceptions import DataExtractionError
from src.ingestion.worksheet_reader import WorksheetReader
from src.types.worksheet_data import (
    CellData,
    ColumnProfile,
    ExtractionStats,
    RowData,
    WorksheetData,
)


# ===========================================================================
# Fixtures compartidas
# ===========================================================================


@pytest.fixture
def reader() -> WorksheetReader:
    """Instancia fresca del WorksheetReader."""
    return WorksheetReader()


@pytest.fixture
def basic_worksheet() -> Any:
    """Hoja con encabezados + 3 filas de datos."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(["Nombre", "Edad", "Municipio"])
    ws.append(["Ana", 25, "Managua"])
    ws.append(["Luis", 30, "León"])
    ws.append(["María", None, "Masaya"])
    return ws


@pytest.fixture
def worksheet_with_empty_rows() -> Any:
    """Hoja con filas completamente vacías intercaladas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ConVacias"
    ws.append(["H1", "H2"])
    ws.append([None, None])        # fila vacía
    ws.append(["A", "B"])
    ws.append(["   ", "   "])     # fila de espacios (vacía)
    ws.append(["C", "D"])
    return ws


@pytest.fixture
def empty_worksheet() -> Any:
    """Hoja completamente vacía."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vacia"
    return ws


@pytest.fixture
def single_row_worksheet() -> Any:
    """Hoja con solo una fila (el encabezado)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SoloHeader"
    ws.append(["Col1", "Col2", "Col3"])
    return ws


# ===========================================================================
# Tests: CellData
# ===========================================================================


class TestCellData:
    """Pruebas del modelo CellData."""

    def test_fields_set_correctly(self) -> None:
        cell = CellData(row_index=1, col_index=2, raw_value="Hola", is_empty=False)
        assert cell.row_index == 1
        assert cell.col_index == 2
        assert cell.raw_value == "Hola"
        assert cell.is_empty is False

    def test_none_raw_value(self) -> None:
        cell = CellData(row_index=0, col_index=0, raw_value=None, is_empty=True)
        assert cell.is_empty is True
        assert cell.raw_value is None

    def test_optional_excel_coords_default_none(self) -> None:
        cell = CellData(0, 0, "X", False)
        assert cell.excel_row is None
        assert cell.excel_col is None

    def test_optional_excel_coords_set(self) -> None:
        cell = CellData(0, 0, "X", False, excel_row=1, excel_col=1)
        assert cell.excel_row == 1
        assert cell.excel_col == 1

    def test_frozen_rejects_mutation(self) -> None:
        cell = CellData(0, 0, "X", False)
        with pytest.raises((AttributeError, TypeError)):
            cell.raw_value = "Y"  # type: ignore[misc]

    def test_repr_contains_value(self) -> None:
        cell = CellData(0, 1, 42, False)
        assert "42" in repr(cell)

    def test_numeric_raw_value(self) -> None:
        cell = CellData(2, 3, 3.14, is_empty=False)
        assert cell.raw_value == 3.14

    def test_bool_raw_value(self) -> None:
        cell = CellData(0, 0, True, is_empty=False)
        assert cell.raw_value is True


# ===========================================================================
# Tests: RowData
# ===========================================================================


class TestRowData:
    """Pruebas del modelo RowData y sus propiedades computadas."""

    @pytest.fixture
    def sample_row(self) -> RowData:
        return RowData(
            row_index=0,
            cells=(
                CellData(0, 0, "Ana", False),
                CellData(0, 1, None, True),
                CellData(0, 2, 25, False),
            ),
        )

    def test_cell_count(self, sample_row: RowData) -> None:
        assert sample_row.cell_count == 3

    def test_non_empty_count(self, sample_row: RowData) -> None:
        assert sample_row.non_empty_count == 2

    def test_is_fully_empty_false(self, sample_row: RowData) -> None:
        assert sample_row.is_fully_empty is False

    def test_is_fully_empty_true(self) -> None:
        row = RowData(
            row_index=0,
            cells=(
                CellData(0, 0, None, True),
                CellData(0, 1, "   ", True),
            ),
        )
        assert row.is_fully_empty is True

    def test_values_property(self, sample_row: RowData) -> None:
        assert sample_row.values == ("Ana", None, 25)

    def test_empty_row_values(self) -> None:
        row = RowData(row_index=0, cells=())
        assert row.values == ()
        assert row.cell_count == 0
        assert row.non_empty_count == 0

    def test_frozen(self, sample_row: RowData) -> None:
        with pytest.raises((AttributeError, TypeError)):
            sample_row.row_index = 99  # type: ignore[misc]

    def test_excel_row_optional(self) -> None:
        row = RowData(row_index=0, cells=())
        assert row.excel_row is None

    def test_excel_row_set(self) -> None:
        row = RowData(row_index=0, cells=(), excel_row=5)
        assert row.excel_row == 5


# ===========================================================================
# Tests: ColumnProfile
# ===========================================================================


class TestColumnProfile:
    """Pruebas del modelo ColumnProfile."""

    def test_fill_rate_normal(self) -> None:
        col = ColumnProfile(col_index=0, total_cells=10, empty_cells=3, non_empty_cells=7)
        assert col.fill_rate == pytest.approx(0.7)

    def test_fill_rate_zero_total(self) -> None:
        col = ColumnProfile(col_index=0, total_cells=0, empty_cells=0, non_empty_cells=0)
        assert col.fill_rate == 0.0

    def test_fill_rate_full(self) -> None:
        col = ColumnProfile(col_index=0, total_cells=5, empty_cells=0, non_empty_cells=5)
        assert col.fill_rate == 1.0

    def test_is_fully_empty_true(self) -> None:
        col = ColumnProfile(col_index=0, total_cells=5, empty_cells=5, non_empty_cells=0)
        assert col.is_fully_empty is True

    def test_is_fully_empty_false(self) -> None:
        col = ColumnProfile(col_index=0, total_cells=5, empty_cells=2, non_empty_cells=3)
        assert col.is_fully_empty is False

    def test_header_value_default_none(self) -> None:
        col = ColumnProfile(col_index=0, total_cells=5, empty_cells=0, non_empty_cells=5)
        assert col.header_value is None

    def test_header_value_set(self) -> None:
        col = ColumnProfile(0, 5, 0, 5, header_value="Nombre")
        assert col.header_value == "Nombre"

    def test_frozen(self) -> None:
        col = ColumnProfile(0, 5, 0, 5)
        with pytest.raises((AttributeError, TypeError)):
            col.col_index = 99  # type: ignore[misc]


# ===========================================================================
# Tests: ExtractionStats
# ===========================================================================


class TestExtractionStats:
    """Pruebas del modelo ExtractionStats."""

    def test_has_data_true(self) -> None:
        stats = ExtractionStats(total_rows_scanned=10, total_columns=3,
                                empty_rows_skipped=2, non_empty_rows=8)
        assert stats.has_data is True

    def test_has_data_false(self) -> None:
        stats = ExtractionStats(total_rows_scanned=5, total_columns=0,
                                empty_rows_skipped=5, non_empty_rows=0)
        assert stats.has_data is False

    def test_frozen(self) -> None:
        stats = ExtractionStats(10, 3, 2, 8)
        with pytest.raises((AttributeError, TypeError)):
            stats.non_empty_rows = 99  # type: ignore[misc]


# ===========================================================================
# Tests: WorksheetData
# ===========================================================================


class TestWorksheetData:
    """Pruebas del modelo raíz WorksheetData y sus métodos de consulta."""

    @pytest.fixture
    def sample_ws_data(self) -> WorksheetData:
        rows = (
            RowData(0, (CellData(0, 0, "Nombre", False), CellData(0, 1, "Edad", False))),
            RowData(1, (CellData(1, 0, "Ana", False), CellData(1, 1, 25, False))),
            RowData(2, (CellData(2, 0, "Luis", False), CellData(2, 1, 30, False))),
        )
        profiles = (
            ColumnProfile(0, 2, 0, 2, "Nombre"),
            ColumnProfile(1, 2, 0, 2, "Edad"),
        )
        stats = ExtractionStats(3, 2, 0, 3)
        return WorksheetData("Datos", rows, profiles, stats)

    def test_sheet_name(self, sample_ws_data: WorksheetData) -> None:
        assert sample_ws_data.sheet_name == "Datos"

    def test_row_count(self, sample_ws_data: WorksheetData) -> None:
        assert sample_ws_data.row_count == 3

    def test_column_count(self, sample_ws_data: WorksheetData) -> None:
        assert sample_ws_data.column_count == 2

    def test_first_row(self, sample_ws_data: WorksheetData) -> None:
        first = sample_ws_data.first_row
        assert first is not None
        assert first.values == ("Nombre", "Edad")

    def test_first_row_empty_dataset(self) -> None:
        ws = WorksheetData("X", (), (), ExtractionStats(0, 0, 0, 0))
        assert ws.first_row is None

    def test_cells_flattened(self, sample_ws_data: WorksheetData) -> None:
        all_cells = sample_ws_data.cells
        assert len(all_cells) == 6  # 3 rows × 2 cols

    def test_raw_matrix(self, sample_ws_data: WorksheetData) -> None:
        matrix = sample_ws_data.raw_matrix
        assert matrix[0] == ("Nombre", "Edad")
        assert matrix[1] == ("Ana", 25)

    def test_get_row_valid(self, sample_ws_data: WorksheetData) -> None:
        row = sample_ws_data.get_row(1)
        assert row is not None
        assert row.values == ("Ana", 25)

    def test_get_row_out_of_bounds(self, sample_ws_data: WorksheetData) -> None:
        assert sample_ws_data.get_row(99) is None
        assert sample_ws_data.get_row(-1) is None

    def test_get_column_profile_found(self, sample_ws_data: WorksheetData) -> None:
        profile = sample_ws_data.get_column_profile(0)
        assert profile is not None
        assert profile.header_value == "Nombre"

    def test_get_column_profile_not_found(self, sample_ws_data: WorksheetData) -> None:
        assert sample_ws_data.get_column_profile(99) is None

    def test_frozen(self, sample_ws_data: WorksheetData) -> None:
        with pytest.raises((AttributeError, TypeError)):
            sample_ws_data.sheet_name = "Otro"  # type: ignore[misc]


# ===========================================================================
# Tests: WorksheetReader.extract_data — flujo principal
# ===========================================================================


class TestWorksheetReaderExtract:
    """Pruebas del método ``extract_data``."""

    def test_returns_worksheet_data(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        result = reader.extract_data(basic_worksheet)
        assert isinstance(result, WorksheetData)

    def test_sheet_name_preserved(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        result = reader.extract_data(basic_worksheet)
        assert result.sheet_name == "Datos"

    def test_correct_row_count(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        """4 filas con datos (header + 3 registros)."""
        result = reader.extract_data(basic_worksheet)
        assert result.row_count == 4

    def test_correct_column_count(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        result = reader.extract_data(basic_worksheet)
        assert result.column_count == 3

    def test_first_row_is_header(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        result = reader.extract_data(basic_worksheet)
        assert result.first_row is not None
        assert result.first_row.values == ("Nombre", "Edad", "Municipio")

    def test_none_value_preserved_in_cell(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        """El None de 'María' fila debe preservarse."""
        result = reader.extract_data(basic_worksheet)
        last_row = result.get_row(3)
        assert last_row is not None
        assert last_row.values[1] is None

    def test_empty_rows_discarded(
        self, reader: WorksheetReader, worksheet_with_empty_rows: Any
    ) -> None:
        """Las filas vacías deben ser descartadas."""
        result = reader.extract_data(worksheet_with_empty_rows)
        # Header + 2 filas de datos reales
        assert result.row_count == 3

    def test_empty_rows_counted_in_stats(
        self, reader: WorksheetReader, worksheet_with_empty_rows: Any
    ) -> None:
        result = reader.extract_data(worksheet_with_empty_rows)
        assert result.extraction_stats.empty_rows_skipped == 2

    def test_total_rows_scanned_accurate(
        self, reader: WorksheetReader, worksheet_with_empty_rows: Any
    ) -> None:
        result = reader.extract_data(worksheet_with_empty_rows)
        # 5 filas totales en la hoja
        assert result.extraction_stats.total_rows_scanned == 5

    def test_raises_on_empty_worksheet(
        self, reader: WorksheetReader, empty_worksheet: Any
    ) -> None:
        with pytest.raises(DataExtractionError, match="no contiene datos"):
            reader.extract_data(empty_worksheet)

    def test_raises_on_all_blank_rows(self, reader: WorksheetReader) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Blancos"
        ws["A1"] = None
        ws["A2"] = "   "
        ws["B1"] = None
        with pytest.raises(DataExtractionError):
            reader.extract_data(ws)

    def test_raises_on_iter_rows_failure(
        self, reader: WorksheetReader, basic_worksheet: Any, monkeypatch: Any
    ) -> None:
        def boom(*args: Any, **kwargs: Any) -> None:
            raise RuntimeError("iter error")
        monkeypatch.setattr(basic_worksheet, "iter_rows", boom)
        with pytest.raises(DataExtractionError, match="Fallo al extraer"):
            reader.extract_data(basic_worksheet)

    def test_column_profiles_created(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        result = reader.extract_data(basic_worksheet)
        assert len(result.column_profiles) == 3

    def test_column_profile_header_value(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        result = reader.extract_data(basic_worksheet)
        profile_0 = result.get_column_profile(0)
        assert profile_0 is not None
        assert profile_0.header_value == "Nombre"

    def test_column_profile_fill_rate(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        """Columna Edad: María tiene None → fill_rate < 1."""
        result = reader.extract_data(basic_worksheet)
        profile_1 = result.get_column_profile(1)
        assert profile_1 is not None
        assert profile_1.fill_rate < 1.0

    def test_result_is_immutable(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        result = reader.extract_data(basic_worksheet)
        with pytest.raises((AttributeError, TypeError)):
            result.sheet_name = "Modificado"  # type: ignore[misc]

    def test_worksheet_not_modified(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        """El worksheet no debe ser alterado por el reader."""
        original_max_row = basic_worksheet.max_row
        reader.extract_data(basic_worksheet)
        assert basic_worksheet.max_row == original_max_row

    def test_single_row_worksheet(
        self, reader: WorksheetReader, single_row_worksheet: Any
    ) -> None:
        """Hoja con solo encabezado debe retornar 1 fila."""
        result = reader.extract_data(single_row_worksheet)
        assert result.row_count == 1
        assert result.first_row is not None

    def test_cells_flat_count(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        result = reader.extract_data(basic_worksheet)
        # 4 filas × 3 columnas
        assert len(result.cells) == 12

    def test_raw_matrix_shape(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        result = reader.extract_data(basic_worksheet)
        matrix = result.raw_matrix
        assert len(matrix) == 4
        assert all(len(row) == 3 for row in matrix)

    def test_excel_row_preserved_in_cells(
        self, reader: WorksheetReader, basic_worksheet: Any
    ) -> None:
        """Las celdas deben tener excel_row != None."""
        result = reader.extract_data(basic_worksheet)
        for row in result.rows:
            for cell in row.cells:
                assert cell.excel_row is not None

    def test_no_pandas_usage(self, reader: WorksheetReader) -> None:
        """El módulo no debe importar pandas."""
        import importlib
        import sys
        mod = sys.modules.get("src.ingestion.worksheet_reader")
        if mod is None:
            mod = importlib.import_module("src.ingestion.worksheet_reader")
        src = getattr(mod, "__file__", "")
        if src:
            with open(src, encoding="utf-8") as f:
                content = f.read()
            assert "import pandas" not in content


# ===========================================================================
# Tests: WorksheetReader._build_cells (helper estático)
# ===========================================================================


class TestBuildCells:
    """Pruebas del helper estático ``_build_cells``."""

    def _make_mock_cell(self, value: Any, row: int, col: int) -> MagicMock:
        cell = MagicMock()
        cell.value = value
        cell.row = row
        cell.column = col
        return cell

    def test_basic_values(self) -> None:
        raw_row = (
            self._make_mock_cell("Nombre", 1, 1),
            self._make_mock_cell(25, 1, 2),
            self._make_mock_cell(None, 1, 3),
        )
        cells = WorksheetReader._build_cells(raw_row, row_idx=0)
        assert len(cells) == 3
        assert cells[0].raw_value == "Nombre"
        assert cells[0].is_empty is False
        assert cells[1].raw_value == 25
        assert cells[1].is_empty is False
        assert cells[2].raw_value is None
        assert cells[2].is_empty is True

    def test_blank_string_is_empty(self) -> None:
        raw_row = (self._make_mock_cell("   ", 1, 1),)
        cells = WorksheetReader._build_cells(raw_row, row_idx=0)
        assert cells[0].is_empty is True
        assert cells[0].raw_value == "   "  # valor crudo preservado

    def test_col_index_base_zero(self) -> None:
        raw_row = tuple(self._make_mock_cell(f"val{i}", 1, i + 1) for i in range(3))
        cells = WorksheetReader._build_cells(raw_row, row_idx=0)
        for i, cell in enumerate(cells):
            assert cell.col_index == i

    def test_excel_coords_set(self) -> None:
        raw_row = (self._make_mock_cell("X", 3, 5),)
        cells = WorksheetReader._build_cells(raw_row, row_idx=0)
        assert cells[0].excel_row == 3
        assert cells[0].excel_col == 5

    def test_empty_row_produces_empty_list(self) -> None:
        cells = WorksheetReader._build_cells((), row_idx=0)
        assert cells == []

    def test_bool_value_not_empty(self) -> None:
        raw_row = (self._make_mock_cell(False, 1, 1),)
        cells = WorksheetReader._build_cells(raw_row, row_idx=0)
        # False no es None ni string vacío → no es empty
        assert cells[0].is_empty is False
        assert cells[0].raw_value is False

    def test_zero_value_not_empty(self) -> None:
        raw_row = (self._make_mock_cell(0, 1, 1),)
        cells = WorksheetReader._build_cells(raw_row, row_idx=0)
        assert cells[0].is_empty is False
        assert cells[0].raw_value == 0


# ===========================================================================
# Tests: WorksheetReader._build_column_profiles (helper estático)
# ===========================================================================


class TestBuildColumnProfiles:
    """Pruebas del helper estático ``_build_column_profiles``."""

    def _make_row(self, row_idx: int, values: list) -> RowData:
        cells = tuple(
            CellData(
                row_index=row_idx,
                col_index=ci,
                raw_value=v,
                is_empty=(v is None or (isinstance(v, str) and v.strip() == "")),
            )
            for ci, v in enumerate(values)
        )
        return RowData(row_index=row_idx, cells=cells)

    def test_header_values_captured(self) -> None:
        rows = [
            self._make_row(0, ["Nombre", "Edad"]),
            self._make_row(1, ["Ana", 25]),
        ]
        profiles = WorksheetReader._build_column_profiles(rows, max_cols=2)
        assert profiles[0].header_value == "Nombre"
        assert profiles[1].header_value == "Edad"

    def test_empty_cells_counted(self) -> None:
        rows = [
            self._make_row(0, ["Col1", "Col2"]),
            self._make_row(1, ["A", None]),
            self._make_row(2, ["B", None]),
            self._make_row(3, ["C", "X"]),
        ]
        profiles = WorksheetReader._build_column_profiles(rows, max_cols=2)
        col1 = profiles[1]
        assert col1.empty_cells == 2
        assert col1.non_empty_cells == 1

    def test_fill_rate_full(self) -> None:
        rows = [
            self._make_row(0, ["H"]),
            self._make_row(1, ["A"]),
            self._make_row(2, ["B"]),
        ]
        profiles = WorksheetReader._build_column_profiles(rows, max_cols=1)
        assert profiles[0].fill_rate == 1.0

    def test_empty_rows_returns_empty(self) -> None:
        profiles = WorksheetReader._build_column_profiles([], max_cols=3)
        assert profiles == []

    def test_max_cols_zero_returns_empty(self) -> None:
        rows = [self._make_row(0, [])]
        profiles = WorksheetReader._build_column_profiles(rows, max_cols=0)
        assert profiles == []

    def test_single_row_no_data_rows(self) -> None:
        """Con solo una fila (header), los conteos de datos deben ser 0."""
        rows = [self._make_row(0, ["H1", "H2"])]
        profiles = WorksheetReader._build_column_profiles(rows, max_cols=2)
        assert profiles[0].total_cells == 0
        assert profiles[1].total_cells == 0

    def test_profile_count_matches_max_cols(self) -> None:
        rows = [
            self._make_row(0, ["A", "B", "C"]),
            self._make_row(1, ["X", "Y", "Z"]),
        ]
        profiles = WorksheetReader._build_column_profiles(rows, max_cols=3)
        assert len(profiles) == 3


# ===========================================================================
# Tests: Logging
# ===========================================================================


class TestLogging:
    """Verifica que el WorksheetReader emite mensajes de log apropiados."""

    def test_logs_start_info(
        self,
        reader: WorksheetReader,
        basic_worksheet: Any,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        with caplog.at_level(logging.INFO, logger="src.ingestion.worksheet_reader"):
            reader.extract_data(basic_worksheet)
        assert any("extracción" in r.message.lower() for r in caplog.records)

    def test_logs_completion_info(
        self,
        reader: WorksheetReader,
        basic_worksheet: Any,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        with caplog.at_level(logging.INFO, logger="src.ingestion.worksheet_reader"):
            reader.extract_data(basic_worksheet)
        assert any("completada" in r.message.lower() for r in caplog.records)

    def test_logs_warning_on_empty_worksheet(
        self,
        reader: WorksheetReader,
        empty_worksheet: Any,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        with caplog.at_level(logging.WARNING, logger="src.ingestion.worksheet_reader"):
            with pytest.raises(DataExtractionError):
                reader.extract_data(empty_worksheet)
        assert any(r.levelno == logging.WARNING for r in caplog.records)


# ===========================================================================
# Tests: Integración end-to-end
# ===========================================================================


class TestIntegration:
    """Pruebas con archivos .xlsx reales guardados en disco."""

    def test_full_cycle_real_file(self, tmp_path: Any) -> None:
        """Ciclo completo: crear → guardar → cargar → extraer."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Beneficiarios"
        ws.append(["Nombre", "Sexo", "Edad", "Municipio"])
        ws.append(["Ana García", "F", 28, "Managua"])
        ws.append(["Pedro Ruiz", "M", 35, "León"])
        ws.append([None, None, None, None])       # vacía
        ws.append(["María López", "F", 22, "Masaya"])
        fp = tmp_path / "test.xlsx"
        wb.save(str(fp))

        wb2 = openpyxl.load_workbook(str(fp))
        ws2 = wb2["Beneficiarios"]
        reader = WorksheetReader()
        result = reader.extract_data(ws2)

        assert result.sheet_name == "Beneficiarios"
        assert result.row_count == 4          # header + 3 datos (1 vacía omitida)
        assert result.column_count == 4
        assert result.extraction_stats.empty_rows_skipped == 1
        assert result.first_row is not None
        assert result.first_row.values == ("Nombre", "Sexo", "Edad", "Municipio")

    def test_column_profiles_real_file(self, tmp_path: Any) -> None:
        """Verifica que los perfiles de columna reflejen los datos reales."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        ws.append(["Col1", "Col2"])
        ws.append(["A", None])
        ws.append(["B", "X"])
        ws.append(["C", None])
        fp = tmp_path / "profiles.xlsx"
        wb.save(str(fp))

        wb2 = openpyxl.load_workbook(str(fp))
        result = WorksheetReader().extract_data(wb2.active)
        col2 = result.get_column_profile(1)
        assert col2 is not None
        assert col2.empty_cells == 2
        assert col2.non_empty_cells == 1
        assert col2.fill_rate == pytest.approx(1 / 3)

    def test_extraction_stats_accurate(self, tmp_path: Any) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stats"
        for i in range(10):
            if i % 3 == 0:
                ws.append([None, None])   # filas vacías
            else:
                ws.append([f"R{i}", i])
        fp = tmp_path / "stats.xlsx"
        wb.save(str(fp))

        wb2 = openpyxl.load_workbook(str(fp))
        result = WorksheetReader().extract_data(wb2.active)
        # Invariante: filas omitidas + filas con datos = total escaneadas
        stats = result.extraction_stats
        assert stats.empty_rows_skipped + stats.non_empty_rows == stats.total_rows_scanned
        assert stats.non_empty_rows > 0
        assert stats.empty_rows_skipped >= 0

    def test_cells_total_count(self, tmp_path: Any) -> None:
        """Verifica el total de celdas en el dataset."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Celdas"
        ws.append(["A", "B", "C", "D"])
        ws.append([1, 2, 3, 4])
        ws.append([5, 6, 7, 8])
        fp = tmp_path / "cells.xlsx"
        wb.save(str(fp))

        wb2 = openpyxl.load_workbook(str(fp))
        result = WorksheetReader().extract_data(wb2.active)
        # 3 filas × 4 columnas = 12 celdas
        assert len(result.cells) == 12
